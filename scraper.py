"""
scraper.py
==========
HBYS automation — reads TC kimlik numbers from an Excel file, finds each
patient in the Poliklinik module via "Hasta Geçmişi", checks for a Yatış
(inpatient admission) record in the current calendar month across all
hospitals, and writes the ward name (Birim) + admission date (Sevk Tarihi)
back to the Excel file.

Workflow per patient:
  1. Enter TC in Kimlik No: field → click Sorgula
  2. Click the patient row in the results grid
  3. Click Hasta Geçmişi button
  4. Click Tüm Hastaneler (red square → all-hospital view)
  5. Scan popup rows for current-month row where Sevk Tipi == "Yatış"
  6. Extract Birim + Sevk Tarihi from that row
  7. Close popup → repeat

Usage:
    python scraper.py

Prerequisites:
    1. pip install -r requirements.txt
    2. Create .env:  HBYS_USERNAME=...  HBYS_PASSWORD=...
    3. Place input.xlsx with TC numbers in column A from row 2
    4. Verify/fill TODO selectors in selectors.py
"""

import sys
import time
import logging
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    StaleElementReferenceException,
)
import openpyxl

import config
import locators as SEL

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Driver
# ---------------------------------------------------------------------------

def make_driver() -> webdriver.Chrome:
    options = webdriver.ChromeOptions()
    # Uncomment for headless (no browser window) production runs:
    # options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    driver = webdriver.Chrome(service=Service(config.CHROMEDRIVER_PATH), options=options)
    driver.implicitly_wait(0)
    return driver

# ---------------------------------------------------------------------------
# Wait helpers
# ---------------------------------------------------------------------------

def wait_visible(driver, by, selector, timeout=None):
    t = timeout or config.WAIT_TIMEOUT
    return WebDriverWait(driver, t).until(
        EC.visibility_of_element_located((by, selector))
    )


def wait_clickable(driver, by, selector, timeout=None):
    t = timeout or config.WAIT_TIMEOUT
    return WebDriverWait(driver, t).until(
        EC.element_to_be_clickable((by, selector))
    )


def safe_find(context, by, selector):
    try:
        return context.find_element(by, selector)
    except NoSuchElementException:
        return None


def switch_to_iframe(driver, css_selector):
    if css_selector is None:
        return
    frame = wait_visible(driver, By.CSS_SELECTOR, css_selector)
    driver.switch_to.frame(frame)

# ---------------------------------------------------------------------------
# Date parsing
# ---------------------------------------------------------------------------

def is_same_month(date_string: str, ref_date: datetime) -> bool:
    """Return True if date_string falls in the same calendar month as ref_date."""
    for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(date_string.strip(), fmt)
            return dt.year == ref_date.year and dt.month == ref_date.month
        except ValueError:
            continue
    log.warning("Could not parse date: '%s'", date_string)
    return False

# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def load_workbook(filepath: str):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    tc_col    = openpyxl.utils.column_index_from_string(config.TC_COLUMN)
    tarih_col = openpyxl.utils.column_index_from_string(config.TARIH_COLUMN)
    min_col   = min(tc_col, tarih_col)
    max_col   = max(tc_col, tarih_col)
    records = []
    for row in ws.iter_rows(min_row=config.DATA_START_ROW, min_col=min_col, max_col=max_col):
        tc_cell    = row[tc_col - min_col]
        tarih_cell = row[tarih_col - min_col]
        if tc_cell.value is not None:
            tc    = str(tc_cell.value).strip().zfill(11)
            tarih = tarih_cell.value  # datetime or string from Excel
            if isinstance(tarih, str):
                for fmt in ("%Y-%m-%d %H:%M:%S", "%d.%m.%Y %H:%M", "%Y-%m-%d"):
                    try:
                        tarih = datetime.strptime(tarih.strip(), fmt)
                        break
                    except ValueError:
                        continue
            records.append((tc_cell.row, tc, tarih))
    log.info("Loaded %d records from %s", len(records), filepath)
    return records, wb, ws


def ensure_output_columns(ws) -> dict:
    header_row = 1
    headers = {
        ws.cell(row=header_row, column=c).value: c
        for c in range(1, ws.max_column + 1)
    }
    col_map = {}
    for col_name in (config.COL_YATIS_VAR, config.COL_BIRIM, config.COL_SEVK_TARIHI):
        if col_name not in headers:
            next_col = ws.max_column + 1
            ws.cell(row=header_row, column=next_col, value=col_name)
            headers[col_name] = next_col
        col_map[col_name] = headers[col_name]
    return col_map


def write_result(ws, row_idx: int, col_map: dict,
                 yatis_var: str, birim: str, tarih: str):
    ws.cell(row=row_idx, column=col_map[config.COL_YATIS_VAR],  value=yatis_var)
    ws.cell(row=row_idx, column=col_map[config.COL_BIRIM],      value=birim)
    ws.cell(row=row_idx, column=col_map[config.COL_SEVK_TARIHI], value=tarih)

# ---------------------------------------------------------------------------
# Login
# ---------------------------------------------------------------------------

def dismiss_kritik_stok_popup(driver):
    """Dismiss the 'Kritik Stok' dialog that appears after every login."""
    try:
        btn = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, SEL.KRITIK_STOK_HAYIR_BUTTON))
        )
        btn.click()
        log.info("Dismissed Kritik Stok popup.")
    except TimeoutException:
        pass  # Popup didn't appear — continue normally


def login(driver):
    log.info("Navigating to %s", config.BASE_URL)
    driver.get(config.BASE_URL)
    switch_to_iframe(driver, SEL.MAIN_APP_IFRAME)

    wait_clickable(driver, By.CSS_SELECTOR, SEL.LOGIN_USERNAME_INPUT).send_keys(config.USERNAME)

    # Click password field — this blurs username and triggers org store load
    password_el = wait_clickable(driver, By.CSS_SELECTOR, SEL.LOGIN_PASSWORD_INPUT)
    password_el.click()

    # Wait for org store to filter to exactly 1 record (user's hospital only)
    # Store starts with 7 items on page load, drops to 1 after username blur
    WebDriverWait(driver, config.WAIT_TIMEOUT).until(
        lambda d: d.execute_script(
            "try { return Ext.ComponentQuery.query('combobox[name=\"kullaniciOrganizasyon\"]')[0]"
            "?.getStore()?.getCount() === 1; } catch(e) { return false; }"
        )
    )

    # Pass the full record to setValue — more reliable than passing value field alone
    org_name = driver.execute_script("""
        var c = Ext.ComponentQuery.query('combobox[name="kullaniciOrganizasyon"]')[0];
        var rec = c.getStore().getAt(0);
        c.setValue(rec);
        c.fireEvent('select', c, [rec]);
        return rec.get(c.displayField);
    """)
    log.info("Selected org: '%s'", org_name)
    time.sleep(0.5)

    # Re-find password field in case element reference went stale
    password_el = wait_clickable(driver, By.CSS_SELECTOR, SEL.LOGIN_PASSWORD_INPUT)
    password_el.clear()
    password_el.send_keys(config.PASSWORD)
    time.sleep(0.3)

    submit = wait_clickable(driver, By.CSS_SELECTOR, SEL.LOGIN_SUBMIT_BUTTON)
    driver.execute_script("arguments[0].click();", submit)

    # Wait for Poliklinik tile on desktop — confirms login completed
    WebDriverWait(driver, config.WAIT_TIMEOUT).until(
        EC.presence_of_element_located((By.XPATH, SEL.POLIKLINIK_MENU_ITEM))
    )
    log.info("Login successful.")
    dismiss_kritik_stok_popup(driver)

# ---------------------------------------------------------------------------
# Open Poliklinik module
# ---------------------------------------------------------------------------

def open_poliklinik_module(driver):
    log.info("Opening Poliklinik module...")
    driver.switch_to.default_content()
    switch_to_iframe(driver, SEL.MAIN_APP_IFRAME)
    tile = wait_clickable(driver, By.XPATH, SEL.POLIKLINIK_MENU_ITEM)
    ActionChains(driver).move_to_element(tile).click().perform()
    time.sleep(3)

    # Click the Poliklinik tab in the module toolbar — without this the panel renders empty
    try:
        tab = wait_clickable(driver, By.XPATH, SEL.POLIKLINIK_TAB, timeout=10)
        tab.click()
        log.info("Clicked Poliklinik tab.")
        time.sleep(1)
    except TimeoutException:
        log.info("Poliklinik tab not found — assuming content loaded directly.")

    switch_to_iframe(driver, SEL.POLIKLINIK_IFRAME)
    wait_visible(driver, By.XPATH, SEL.POLIKLINIK_LOADED_LANDMARK, timeout=30)
    log.info("Poliklinik module loaded.")

    # Widen date range so March patients are visible
    driver.execute_script("""
        try {
            var fields = Ext.ComponentQuery.query('datefield');
            var now = new Date();
            var firstOfLastMonth = new Date(now.getFullYear(), now.getMonth() - 1, 1);
            for (var i = 0; i < fields.length; i++) {
                var lbl = (fields[i].fieldLabel || '').trim();
                if (lbl.indexOf('lk Tarih') !== -1) {
                    fields[i].setValue(firstOfLastMonth);
                } else if (lbl === 'Son Tarih:' || lbl === 'Son Tarih') {
                    fields[i].setValue(now);
                }
            }
        } catch(e) {}
    """)
    log.info("Date range set: 1st of last month → today.")

# ---------------------------------------------------------------------------
# Search patient by TC
# ---------------------------------------------------------------------------

def search_patient(driver, tc: str) -> bool:
    """Enter TC and click Sorgula. Returns True if at least one row appears."""
    log.info("Searching TC: %s", tc)

    # Clear via Temizle first to reset grid state
    try:
        temizle = driver.find_element(By.XPATH, SEL.CLEAR_SEARCH_BUTTON)
        driver.execute_script("arguments[0].click();", temizle)
        time.sleep(1)
    except NoSuchElementException:
        pass

    tc_input = wait_clickable(driver, By.XPATH, SEL.TC_SEARCH_INPUT)
    tc_input.clear()
    time.sleep(0.5)
    tc_input.send_keys(tc)
    time.sleep(0.5)

    btn = wait_clickable(driver, By.XPATH, SEL.TC_SEARCH_BUTTON)
    driver.execute_script("arguments[0].click();", btn)

    # Wait for the grid to load — the grid shows a loading mask during search
    time.sleep(3)  # give the slow system time to start the request

    try:
        WebDriverWait(driver, config.WAIT_TIMEOUT).until(
            EC.presence_of_element_located((By.XPATH, SEL.PATIENT_ROW_XPATH))
        )
    except TimeoutException:
        log.info("No patient rows found for TC: %s", tc)
        return False

    # Extra settle time for grid to fully render
    time.sleep(1)
    rows = driver.find_elements(By.XPATH, SEL.PATIENT_ROW_XPATH)
    log.info("Found %d row(s) for TC: %s", len(rows), tc)
    return bool(rows)

# ---------------------------------------------------------------------------
# Select patient row closest to transfer date
# ---------------------------------------------------------------------------

def select_patient_row(driver):
    """
    Click the first visible patient row. Simply selects the first row to
    open the patient — the Hasta Geçmişi popup searches all visits anyway.
    """
    rows = driver.find_elements(By.XPATH, SEL.PATIENT_ROW_XPATH)
    if not rows:
        raise RuntimeError("No patient rows to select.")

    log.info("Clicking first patient row out of %d.", len(rows))
    driver.execute_script("arguments[0].click();", rows[0])
    time.sleep(config.NAV_SETTLE_WAIT)

# ---------------------------------------------------------------------------
# Open Hasta Geçmişi popup
# ---------------------------------------------------------------------------

def open_hasta_gecmisi(driver):
    """Click the Hasta Geçmişi button and wait for the popup."""
    log.info("Opening Hasta Geçmişi...")
    btn = wait_clickable(driver, By.XPATH, SEL.HASTA_GECMISI_BUTTON)
    driver.execute_script("arguments[0].click();", btn)
    wait_visible(driver, By.XPATH, SEL.GECMIS_POPUP)
    time.sleep(1)  # let popup fully render
    log.info("Hasta Geçmişi popup opened.")

# ---------------------------------------------------------------------------
# Click Tüm Hastaneler
# ---------------------------------------------------------------------------

def click_tum_hastaneler(driver):
    """
    Click the red 'Tüm Hastaneler' toggle inside the popup to load
    results from all connected hospitals (including the tertiary centre).
    After clicking, the toggle turns green and the grid reloads.
    TUM_HASTANELER_BUTTON is a relative XPath — scoped to the popup element.
    """
    log.info("Clicking Tüm Hastaneler...")
    wait_visible(driver, By.XPATH, SEL.GECMIS_POPUP)
    time.sleep(1)  # allow popup grid to fully render before interacting

    # Use ExtJS component API to toggle the checkbox — DOM click stops working
    # after the first popup because ExtJS doesn't rebind the event handler
    clicked = driver.execute_script("""
        // Find the Hasta Geçmişi popup
        var wins = Ext.ComponentQuery.query('window');
        var popup = null;
        for (var i = 0; i < wins.length; i++) {
            if (wins[i].title && wins[i].title.indexOf('Hasta Ge') !== -1 && wins[i].isVisible()) {
                popup = wins[i];
                break;
            }
        }
        if (!popup) return 'popup not found';

        // Find the checkbox inside the popup via ComponentQuery
        var cbs = popup.query('checkboxfield');
        var cb = null;
        for (var i = 0; i < cbs.length; i++) {
            var lbl = cbs[i].fieldLabel || '';
            if (lbl.indexOf('Hastaneler') !== -1) { cb = cbs[i]; break; }
        }
        if (!cb) return 'checkbox not found in popup (found ' + cbs.length + ' checkboxes)';

        // If already checked, uncheck first to force a fresh load
        if (cb.getValue()) {
            cb.setValue(false);
            cb.fireEvent('change', cb, false, true);
        }

        // Small delay handled by setTimeout, then check
        cb.setValue(true);
        cb.fireEvent('change', cb, true, false);
        return 'toggled via API: ' + cb.id;
    """)
    log.info("Tüm Hastaneler JS result: %s", clicked)
    time.sleep(5)  # wait for grid to reload with all-hospital data
    log.info("Tüm Hastaneler loaded.")

# ---------------------------------------------------------------------------
# Scan popup rows for Yatış
# ---------------------------------------------------------------------------

def find_yatis_in_popup(driver, transfer_dt: datetime) -> dict:
    """
    Scan all rows in the Hasta Geçmişi popup using the ExtJS grid store.
    Look for a row where yatisTarihi is set (indicates internment) and falls
    in the same month as the transfer date.
    Returns {"found": bool, "birim": str, "tarih": str}.
    """
    result = {"found": False, "birim": "", "tarih": ""}
    wait_visible(driver, By.XPATH, SEL.GECMIS_POPUP)

    target_month = transfer_dt.month
    target_year  = transfer_dt.year

    data = driver.execute_script("""
        var targetMonth = arguments[0];
        var targetYear  = arguments[1];

        var wins = Ext.ComponentQuery.query('window');
        var popup = null;
        for (var i = 0; i < wins.length; i++) {
            if (wins[i].title && wins[i].title.indexOf('Hasta Ge') !== -1 && wins[i].isVisible()) {
                popup = wins[i];
                break;
            }
        }
        if (!popup) return {error: 'popup not found'};

        var grids = popup.query('gridpanel');
        if (!grids.length) return {error: 'no grid in popup'};

        var store = grids[0].getStore();
        if (!store) return {error: 'no store'};

        var count = store.getCount();
        var matches = [];
        var sevkTipiValues = {};

        for (var i = 0; i < count; i++) {
            var rec = store.getAt(i);
            var d = rec.data;
            var st = String(d.sevkTipi || '');
            sevkTipiValues[st] = (sevkTipiValues[st] || 0) + 1;

            // sevkTipi '2' = Yatış (inpatient admission)
            if (st !== '2') continue;

            // Only consider records from the tertiary center
            var org = String(d.birimOrganizasyon || '');
            if (org.indexOf('Ara') === -1) continue;  // must contain "Araştırma"

            // Parse sevkTarihi to check month
            var sevkStr = String(d.sevkTarihi || '');
            var sevkDate = new Date(sevkStr);
            if (isNaN(sevkDate.getTime())) continue;

            if (sevkDate.getMonth() + 1 === targetMonth && sevkDate.getFullYear() === targetYear) {
                matches.push({
                    birim: d.birimAdi || '',
                    tarih: sevkDate.toLocaleDateString('tr-TR'),
                    org: org
                });
            }
        }

        return {count: count, sevkTipiValues: sevkTipiValues, matches: matches};
    """, target_month, target_year)

    if data.get('error'):
        log.warning("Popup store error: %s", data['error'])
        return result

    log.info("Popup: %d rows, sevkTipi values: %s, matches: %d",
             data.get('count', 0), data.get('sevkTipiValues', {}), len(data.get('matches', [])))

    matches = data.get('matches', [])
    if matches:
        m = matches[0]
        log.info("  Yatış found: Birim=%s | Tarih=%s | Org=%s", m['birim'], m['tarih'], m['org'])
        result = {"found": True, "birim": m['birim'], "tarih": m['tarih']}

    return result

# ---------------------------------------------------------------------------
# Close popup
# ---------------------------------------------------------------------------

def close_all_popups(driver):
    """Close only Hasta Geçmişi windows and remove masks — preserves Poliklinik."""
    driver.execute_script("""
        try {
            var wins = Ext.ComponentQuery.query('window{isVisible()}');
            for (var i = wins.length - 1; i >= 0; i--) {
                var t = wins[i].title || '';
                if (t.indexOf('Hasta Ge') !== -1 || t.indexOf('Birim') !== -1) {
                    try { wins[i].close(); } catch(e) {}
                }
            }
        } catch(e) {}
        // Remove any leftover masks
        document.querySelectorAll('.x-mask').forEach(function(m) { m.remove(); });
    """)
    time.sleep(1)


def close_gecmis_popup(driver):
    # Try closing via ExtJS first — targeted to Hasta Geçmişi window only
    closed = driver.execute_script("""
        try {
            var wins = Ext.ComponentQuery.query('window{isVisible()}');
            for (var i = wins.length - 1; i >= 0; i--) {
                var t = wins[i].title || '';
                if (t.indexOf('Hasta Ge') !== -1) {
                    wins[i].close();
                    return true;
                }
            }
        } catch(e) {}
        return false;
    """)

    if closed:
        time.sleep(1)
        log.info("Popup closed via ExtJS.")
    else:
        # Fallback: Escape key
        try:
            webdriver.ActionChains(driver).send_keys(Keys.ESCAPE).perform()
            time.sleep(1)
        except Exception:
            pass
        log.info("Popup closed via Escape fallback.")

    # ALWAYS remove leftover masks — ExtJS close often leaves them behind
    driver.execute_script("""
        document.querySelectorAll('.x-mask').forEach(function(m) { m.remove(); });
    """)

# ---------------------------------------------------------------------------
# Reset search
# ---------------------------------------------------------------------------

def reset_search(driver):
    """Clear Kriter Paneli so the next TC can be entered."""
    try:
        btn = wait_clickable(driver, By.XPATH, SEL.CLEAR_SEARCH_BUTTON, timeout=5)
        driver.execute_script("arguments[0].click();", btn)
    except TimeoutException:
        tc_input = safe_find(driver, By.XPATH, SEL.TC_SEARCH_INPUT)
        if tc_input:
            tc_input.clear()

# ---------------------------------------------------------------------------
# Per-patient pipeline
# ---------------------------------------------------------------------------

def process_patient(driver, tc: str, transfer_dt: datetime) -> dict:
    not_found = {"yatis_var": "Hasta Bulunamadı", "birim": "", "tarih": ""}
    no_yatis  = {"yatis_var": "Hayır",            "birim": "", "tarih": ""}

    if not search_patient(driver, tc):
        return not_found

    try:
        select_patient_row(driver)
    except RuntimeError as exc:
        log.warning("Could not select patient row: %s", exc)
        reset_search(driver)
        return not_found

    open_hasta_gecmisi(driver)
    click_tum_hastaneler(driver)

    info = find_yatis_in_popup(driver, transfer_dt)
    close_gecmis_popup(driver)
    reset_search(driver)

    if info["found"]:
        return {"yatis_var": "Evet", "birim": info["birim"], "tarih": info["tarih"]}
    return no_yatis

# ---------------------------------------------------------------------------
# Health check
# ---------------------------------------------------------------------------

def _poliklinik_is_alive(driver) -> bool:
    """Return True if the Poliklinik search panel is currently visible."""
    try:
        els = driver.find_elements(By.XPATH, SEL.POLIKLINIK_LOADED_LANDMARK)
        return bool(els) and els[0].is_displayed()
    except Exception:
        return False

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if not config.USERNAME or not config.PASSWORD:
        log.error(
            "Credentials not set. Create .env with:\n"
            "  HBYS_USERNAME=...\n  HBYS_PASSWORD=..."
        )
        sys.exit(1)

    records, wb, ws = load_workbook(config.INPUT_FILE)
    col_map = ensure_output_columns(ws)

    driver = make_driver()
    try:
        login(driver)
        open_poliklinik_module(driver)

        for idx, (row_idx, tc, transfer_dt) in enumerate(records, start=1):
            log.info("=== %d / %d | TC: %s ===", idx, len(records), tc)

            # Ensure Poliklinik module is alive before each patient
            if not _poliklinik_is_alive(driver):
                log.warning("Poliklinik module lost — re-opening.")
                try:
                    open_poliklinik_module(driver)
                except Exception:
                    log.exception("Failed to re-open Poliklinik — aborting.")
                    break

            try:
                result = process_patient(driver, tc, transfer_dt)
                write_result(ws, row_idx, col_map,
                             yatis_var=result["yatis_var"],
                             birim=result["birim"],
                             tarih=result["tarih"])
                wb.save(config.INPUT_FILE)
                log.info("Written: %s", result)
            except Exception:
                log.exception("Unhandled error for TC %s (row %d)", tc, row_idx)
                write_result(ws, row_idx, col_map,
                             yatis_var="HATA", birim="", tarih="")
                wb.save(config.INPUT_FILE)
                # Cleanup: close all popups/masks so next patient starts clean
                close_all_popups(driver)

    finally:
        driver.quit()
        log.info("Scraping done.")

    generate_report(ws, col_map)


def generate_report(ws, col_map):
    """Generate a clean report Excel with coloring and tally."""
    from openpyxl.styles import PatternFill, Font, Alignment
    from collections import Counter

    log.info("Generating report: %s", config.OUTPUT_FILE)

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Sevk Raporu"

    # Column indices in input sheet (1-based)
    COL_ADI_SOYADI = 1        # A: ADI_SOYADI
    COL_TARIH = 7             # G: TARIH
    COL_TUMSEVKTANILARI = 8   # H: TUMSEVKTANILARI

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    brown_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

    # Write headers
    headers = ["Tarih", "Hasta Adı", "Sevk Tanısı", "Yatış Var", "Birim"]
    for c, h in enumerate(headers, 1):
        cell = out_ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Collect data and birim tally
    birim_counter = Counter()
    out_row = 2

    for row_idx in range(config.DATA_START_ROW, ws.max_row + 1):
        yatis_var = ws.cell(row=row_idx, column=col_map[config.COL_YATIS_VAR]).value
        if yatis_var is None:
            continue

        tarih = ws.cell(row=row_idx, column=COL_TARIH).value
        adi_soyadi = ws.cell(row=row_idx, column=COL_ADI_SOYADI).value
        sevk_tanisi = ws.cell(row=row_idx, column=COL_TUMSEVKTANILARI).value
        birim = ws.cell(row=row_idx, column=col_map[config.COL_BIRIM]).value or ""

        # Format tarih
        if isinstance(tarih, datetime):
            tarih_str = tarih.strftime("%d.%m.%Y %H:%M")
        else:
            tarih_str = str(tarih) if tarih else ""

        out_ws.cell(row=out_row, column=1, value=tarih_str)
        out_ws.cell(row=out_row, column=2, value=adi_soyadi)
        out_ws.cell(row=out_row, column=3, value=sevk_tanisi)
        out_ws.cell(row=out_row, column=4, value=yatis_var)
        out_ws.cell(row=out_row, column=5, value=birim)

        # Color the row
        if yatis_var == "Evet":
            row_fill = green_fill
            birim_counter[birim] += 1
        elif yatis_var == "Hayır":
            row_fill = brown_fill
        else:
            row_fill = None

        if row_fill:
            for c in range(1, 6):
                out_ws.cell(row=out_row, column=c).fill = row_fill

        out_row += 1

    # Write tally to the right side
    tally_col = 7  # Column G
    tally_header = out_ws.cell(row=1, column=tally_col, value="Yatış Birimi")
    tally_header.fill = header_fill
    tally_header.font = header_font
    count_header = out_ws.cell(row=1, column=tally_col + 1, value="Sayı")
    count_header.fill = header_fill
    count_header.font = header_font

    for i, (birim, count) in enumerate(birim_counter.most_common(), start=2):
        out_ws.cell(row=i, column=tally_col, value=birim)
        out_ws.cell(row=i, column=tally_col + 1, value=count)

    # Auto-width columns
    for col in out_ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        out_ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    out_wb.save(config.OUTPUT_FILE)
    log.info("Report saved: %s", config.OUTPUT_FILE)


if __name__ == "__main__":
    main()
